import sys
import requests
import openpyxl
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                             QFileDialog, QProgressBar, QLabel, QComboBox, QStyleFactory, QMessageBox)
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from PyQt5.QtGui import QIcon
import re
import time

class LinkChecker(QThread):
    update_progress = pyqtSignal(int, int, int, int, int)
    finished = pyqtSignal()
    error = pyqtSignal(str)
    connection_lost = pyqtSignal()
    connection_restored = pyqtSignal()

    def __init__(self, file_path, sheet_name, column):
        super().__init__()
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.column = column
        self.is_paused = False
        self.is_canceled = False

    def is_valid_url(self, url):
        if not url:
            return False
        url = url.strip().replace(" ", "%20")
        regex = re.compile(
            r'^(?:http|ftp)s?://'  # http:// ou https://
            r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|'  # domaine
            r'localhost|'  # localhost
            r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})'  # ou IP
            r'(?::\d+)?'  # port optionnel
            r'(?:/?|[/?]\S+)$', re.IGNORECASE)
        return re.match(regex, str(url)) is not None

    def run(self):
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook[self.sheet_name]

            # Nouvelle colonne pour les statuts
            new_column = get_column_letter(sheet.max_column + 1)
            sheet[f"{new_column}1"] = "Statut du lien"

            total_links = sum(1 for cell in sheet[self.column][1:] if self.is_valid_url(cell.value))
            processed_links = 0
            valid_links = 0
            invalid_links = 0

            for row, cell in enumerate(sheet[self.column][1:], start=2):
                if self.is_canceled:
                    break

                while self.is_paused:
                    time.sleep(1)

                cell_value = cell.value
                if not cell_value or not self.is_valid_url(cell_value):
                    sheet[f"{new_column}{row}"] = "Invalide (URL non valide)"
                    invalid_links += 1
                    processed_links += 1
                    continue

                try:
                    response = requests.head(str(cell_value), timeout=5, allow_redirects=True)
                    if response.status_code == 200:
                        valid_links += 1
                        sheet[f"{new_column}{row}"] = "Valide"
                    else:
                        invalid_links += 1
                        sheet[f"{new_column}{row}"] = f"Invalide (Code: {response.status_code})"
                except requests.RequestException as e:
                    invalid_links += 1
                    sheet[f"{new_column}{row}"] = f"Invalide (Erreur: {str(e)})"

                processed_links += 1
                percentage = int(processed_links / total_links * 100) if total_links > 0 else 0
                self.update_progress.emit(total_links, processed_links, valid_links, invalid_links, percentage)

            workbook.save(self.file_path)
            self.finished.emit()

        except Exception as e:
            self.error.emit(str(e))

class App(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Link Checker XLSX')
        self.setWindowIcon(QIcon('D:/Projets/Python/Python Checklink/LinkChecker.png'))
        self.setGeometry(100, 100, 600, 400)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        layout = QVBoxLayout()
        central_widget.setLayout(layout)

        # Sélection du fichier
        file_layout = QHBoxLayout()
        self.file_label = QLabel('Aucun fichier sélectionné')
        self.file_btn = QPushButton('Choisir un fichier Excel')
        self.file_btn.clicked.connect(self.choose_file)
        file_layout.addWidget(self.file_label)
        file_layout.addWidget(self.file_btn)
        layout.addLayout(file_layout)

        # Sélection de la feuille
        sheet_layout = QHBoxLayout()
        sheet_layout.addWidget(QLabel('Feuille:'))
        self.sheet_combo = QComboBox()
        sheet_layout.addWidget(self.sheet_combo)
        layout.addLayout(sheet_layout)

        # Sélection de la colonne
        column_layout = QHBoxLayout()
        column_layout.addWidget(QLabel('Colonne:'))
        self.column_combo = QComboBox()
        column_layout.addWidget(self.column_combo)
        layout.addLayout(column_layout)

        # Bouton de démarrage
        self.start_btn = QPushButton('Commencer la vérification')
        self.start_btn.clicked.connect(self.start_checking)
        layout.addWidget(self.start_btn)

        # Bouton pause et annulation
        self.pause_btn = QPushButton('Pause')
        self.pause_btn.clicked.connect(self.toggle_pause)
        layout.addWidget(self.pause_btn)
        self.pause_btn.hide()

        self.cancel_btn = QPushButton('Annuler')
        self.cancel_btn.clicked.connect(self.cancel_checking)
        layout.addWidget(self.cancel_btn)
        self.cancel_btn.hide()

        # Barre de progression
        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)

        # Label de statut
        self.status_label = QLabel()
        self.status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.status_label)

        # Label de connexion
        self.connection_label = QLabel()
        self.connection_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.connection_label)

        # Set the style
        self.setStyle(QStyleFactory.create('Fusion'))
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f0f0f0;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QLabel {
                font-size: 14px;
            }
            QComboBox {
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 3px;
            }
        """)

        self.show()

    def choose_file(self):
        if hasattr(self, 'link_checker') and self.link_checker.isRunning():
            reply = QMessageBox.question(self, "Confirmation", 
                                        "Une vérification est en cours. Voulez-vous l'annuler et sélectionner un nouveau fichier ?", 
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                return
            else:
                self.link_checker.is_canceled = True
                self.link_checker.wait()  # Attendre que le thread soit terminé avant de continuer
                del self.link_checker  # Supprimer l'instance du thread

        # Réinitialiser l'interface utilisateur
        self.sheet_combo.clear()
        self.column_combo.clear()
        self.progress_bar.setValue(0)
        self.status_label.setText("")
        self.pause_btn.hide()
        self.cancel_btn.hide()
        self.start_btn.show()

        # Sélectionner un nouveau fichier
        file_path, _ = QFileDialog.getOpenFileName(self, "Choisir un fichier Excel", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            self.file_path = file_path
            self.file_label.setText(f"Fichier sélectionné: {file_path.split('/')[-1]}")
            self.load_sheets()

    def load_sheets(self):
        self.sheet_combo.clear()
        self.column_combo.clear()
        self.progress_bar.setValue(0)
        self.status_label.setText("")
        
        workbook = openpyxl.load_workbook(self.file_path)
        self.sheet_combo.addItems(workbook.sheetnames)
        self.sheet_combo.currentIndexChanged.connect(self.load_columns)
        self.load_columns()

    def load_columns(self):
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook[self.sheet_combo.currentText()]
        self.column_combo.clear()
        self.progress_bar.setValue(0)
        self.status_label.setText("")
        for col in sheet.iter_cols(1, sheet.max_column):
            col_letter = get_column_letter(col[0].column)
            col_title = sheet.cell(row=1, column=col[0].column).value
            if col_title:
                self.column_combo.addItem(f"{col_title} ({col_letter})")
            else:
                self.column_combo.addItem(f"Colonne {col_letter}")

    def start_checking(self):
        self.start_btn.hide()
        self.pause_btn.show()
        self.cancel_btn.show()

        sheet_name = self.sheet_combo.currentText()
        column = get_column_letter(self.column_combo.currentIndex() + 1)
        self.link_checker = LinkChecker(self.file_path, sheet_name, column)
        self.link_checker.update_progress.connect(self.update_progress)
        self.link_checker.finished.connect(self.checking_finished)
        self.link_checker.error.connect(self.display_error)
        self.link_checker.connection_lost.connect(self.connection_lost)
        self.link_checker.connection_restored.connect(self.connection_restored)
        self.link_checker.start()

    def update_progress(self, total, processed, valid, invalid, percentage):
        self.progress_bar.setValue(percentage)
        self.status_label.setText(f"Total: {total}, Traités: {processed}, Valides: {valid}, Invalides: {invalid}")

    def checking_finished(self):
        self.start_btn.show()
        self.pause_btn.hide()
        self.cancel_btn.hide()
        self.status_label.setText("Vérification terminée!")

    def display_error(self, message):
        QMessageBox.critical(self, "Erreur", message)
        self.start_btn.show()
        self.pause_btn.hide()
        self.cancel_btn.hide()

    def toggle_pause(self):
        if self.link_checker.is_paused:
            self.link_checker.is_paused = False
            self.pause_btn.setText('Pause')
        else:
            self.link_checker.is_paused = True
            self.pause_btn.setText('Reprendre')

    def cancel_checking(self):
        reply = QMessageBox.question(self, "Confirmation", "Voulez-vous vraiment annuler la vérification ?", 
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.link_checker.is_canceled = True
            self.pause_btn.hide()
            self.cancel_btn.hide()
            self.start_btn.show()

    def connection_lost(self):
        self.connection_label.setText("Connexion perdue, attente de restauration...")

    def connection_restored(self):
        self.connection_label.setText("Connexion restaurée, reprise en cours...")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = App()
    sys.exit(app.exec_())
s